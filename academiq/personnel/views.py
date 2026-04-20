from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import Group
from django.db import IntegrityError
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.db.models import Avg

import io
from datetime import date
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from core.permissions import role_required
from core.models import (
    AnneeScolaire, Periode, Salle, Matiere, Classe, Cours,
    Personne, Personnel, Enseignant, Eleve, Parent, Inscription,
    Absence, Bulletin, ResultatMatiere, Notification, HistoriqueActions,
    EmploiDuTemps, Message, EvenementScolaire, FraisScolarite, Paiement,
)
from .forms import (
    AnneeScolaireForm, PeriodeForm, SalleForm, MatiereForm, ClasseForm,
    CoursForm, PersonneBaseForm, PersonnelSubForm, EnseignantSubForm,
    EleveSubForm, ParentSubForm, RoleChoiceForm, InscriptionForm,
    EmploiDuTempsForm,
)

ROLES_ADMIN = ('DIRECTION', 'ADMINISTRATION', 'SCOLARITE', 'FINANCES')


# ─── Dashboard ────────────────────────────────────────────────────────────────

@role_required(*ROLES_ADMIN)
def dashboard(request):
    import json
    annee_active = AnneeScolaire.objects.filter(active=True).first()

    # Élèves par classe
    classes = Classe.objects.filter(annee=annee_active) if annee_active else Classe.objects.none()
    eleves_par_classe_labels = []
    eleves_par_classe_data = []
    for cls in classes:
        nb = Inscription.objects.filter(classe=cls, annee=annee_active, statut='actif').count()
        eleves_par_classe_labels.append(cls.nom)
        eleves_par_classe_data.append(nb)

    # Statut des absences
    nb_attente   = Absence.objects.filter(statut='en_attente').count()
    nb_justifie  = Absence.objects.filter(statut='justifiee').count()
    nb_njustifie = Absence.objects.filter(statut='non_justifiee').count()

    # Moyennes par matière (sur toutes les périodes)
    from django.db.models import Avg
    matieres_qs = Matiere.objects.all()
    moy_labels, moy_data = [], []
    for mat in matieres_qs:
        moy = ResultatMatiere.objects.filter(
            cours__matiere=mat
        ).aggregate(m=Avg('moyenne'))['m']
        if moy is not None:
            moy_labels.append(mat.nom_matiere[:20])
            moy_data.append(round(float(moy), 2))

    ctx = {
        'annee_active':    annee_active,
        'nb_eleves':       Personne.objects.filter(eleve__isnull=False, actif=True).count(),
        'nb_enseignants':  Personne.objects.filter(enseignant__isnull=False, actif=True).count(),
        'nb_classes':      classes.count(),
        'nb_cours':        Cours.objects.filter(annee=annee_active).count() if annee_active else 0,
        'absences_attente': nb_attente,
        # Données graphiques (JSON)
        'chart_classes_labels': json.dumps(eleves_par_classe_labels),
        'chart_classes_data':   json.dumps(eleves_par_classe_data),
        'chart_absences_data':  json.dumps([nb_attente, nb_justifie, nb_njustifie]),
        'chart_moy_labels':     json.dumps(moy_labels),
        'chart_moy_data':       json.dumps(moy_data),
    }
    return render(request, 'personnel/dashboard.html', ctx)


# ─── Années scolaires ─────────────────────────────────────────────────────────

@role_required(*ROLES_ADMIN)
def gestion_annees(request):
    annees = AnneeScolaire.objects.all()
    form = AnneeScolaireForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Année scolaire créée avec succès.")
        return redirect('personnel:gestion_annees')
    return render(request, 'personnel/annees/liste.html', {'annees': annees, 'form': form})


@role_required('DIRECTION')
def activer_annee(request, pk):
    annee = get_object_or_404(AnneeScolaire, pk=pk)
    annee.active = True
    annee.save()
    messages.success(request, f"Année {annee.libelle} activée.")
    return redirect('personnel:gestion_annees')


@role_required('DIRECTION')
def modifier_annee(request, pk):
    annee = get_object_or_404(AnneeScolaire, pk=pk)
    form = AnneeScolaireForm(request.POST or None, instance=annee)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Année modifiée.")
        return redirect('personnel:gestion_annees')
    return render(request, 'personnel/annees/form.html', {'form': form, 'objet': annee, 'titre': "Modifier l'année"})


# ─── Périodes ─────────────────────────────────────────────────────────────────

@role_required(*ROLES_ADMIN)
def gestion_periodes(request):
    annee_active = AnneeScolaire.objects.filter(active=True).first()
    periodes = Periode.objects.filter(annee=annee_active) if annee_active else Periode.objects.none()
    form = PeriodeForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Période créée.")
        return redirect('personnel:gestion_periodes')
    return render(request, 'personnel/periodes/liste.html', {
        'periodes': periodes, 'form': form, 'annee_active': annee_active
    })


@role_required('DIRECTION', 'SCOLARITE')
def cloturer_periode(request, pk):
    periode = get_object_or_404(Periode, pk=pk)
    periode.cloturee = True
    periode.save()
    messages.success(request, f"Période «{periode.nom}» clôturée.")
    return redirect('personnel:gestion_periodes')


# ─── Salles ───────────────────────────────────────────────────────────────────

@role_required(*ROLES_ADMIN)
def gestion_salles(request):
    salles = Salle.objects.all()
    form = SalleForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Salle créée.")
        return redirect('personnel:gestion_salles')
    return render(request, 'personnel/salles/liste.html', {'salles': salles, 'form': form})


@role_required(*ROLES_ADMIN)
def modifier_salle(request, pk):
    salle = get_object_or_404(Salle, pk=pk)
    form = SalleForm(request.POST or None, instance=salle)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Salle modifiée.")
        return redirect('personnel:gestion_salles')
    return render(request, 'personnel/salles/form.html', {'form': form, 'objet': salle, 'titre': 'Modifier la salle'})


@role_required('DIRECTION')
def supprimer_salle(request, pk):
    salle = get_object_or_404(Salle, pk=pk)
    if request.method == 'POST':
        salle.delete()
        messages.success(request, "Salle supprimée.")
        return redirect('personnel:gestion_salles')
    return render(request, 'personnel/confirm_suppression.html', {'objet': salle, 'type': 'la salle'})


# ─── Matières ─────────────────────────────────────────────────────────────────

@role_required(*ROLES_ADMIN)
def gestion_matieres(request):
    matieres = Matiere.objects.all()
    form = MatiereForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Matière créée.")
        return redirect('personnel:gestion_matieres')
    return render(request, 'personnel/matieres/liste.html', {'matieres': matieres, 'form': form})


@role_required(*ROLES_ADMIN)
def modifier_matiere(request, pk):
    matiere = get_object_or_404(Matiere, pk=pk)
    form = MatiereForm(request.POST or None, instance=matiere)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Matière modifiée.")
        return redirect('personnel:gestion_matieres')
    return render(request, 'personnel/matieres/form.html', {'form': form, 'objet': matiere, 'titre': 'Modifier la matière'})


# ─── Classes ──────────────────────────────────────────────────────────────────

@role_required('DIRECTION', 'SCOLARITE', 'ADMINISTRATION')
def gestion_classes(request):
    annee_active = AnneeScolaire.objects.filter(active=True).first()
    classes = Classe.objects.filter(annee=annee_active).select_related('annee') if annee_active else Classe.objects.none()
    form = ClasseForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        try:
            form.save()
            messages.success(request, "Classe créée.")
            return redirect('personnel:gestion_classes')
        except IntegrityError:
            messages.error(request, "Cette classe existe déjà pour cette année.")
    return render(request, 'personnel/classes/liste.html', {
        'classes': classes, 'form': form, 'annee_active': annee_active
    })


@role_required('DIRECTION', 'SCOLARITE', 'ADMINISTRATION')
def modifier_classe(request, pk):
    classe = get_object_or_404(Classe, pk=pk)
    form = ClasseForm(request.POST or None, instance=classe)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Classe modifiée.")
        return redirect('personnel:gestion_classes')
    return render(request, 'personnel/classes/form.html', {'form': form, 'objet': classe, 'titre': 'Modifier la classe'})


# ─── Cours ────────────────────────────────────────────────────────────────────

@role_required('DIRECTION', 'ADMINISTRATION')
def gestion_cours(request):
    annee_active = AnneeScolaire.objects.filter(active=True).first()
    cours_qs = Cours.objects.filter(annee=annee_active).select_related(
        'matiere', 'classe', 'enseignant', 'annee'
    ) if annee_active else Cours.objects.none()
    form = CoursForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        try:
            form.save()
            messages.success(request, "Cours créé.")
            return redirect('personnel:gestion_cours')
        except IntegrityError:
            messages.error(request, "Ce cours existe déjà pour cette matière, classe et année.")
    return render(request, 'personnel/cours/liste.html', {
        'cours_list': cours_qs, 'form': form, 'annee_active': annee_active
    })


@role_required('DIRECTION', 'ADMINISTRATION')
def modifier_cours(request, pk):
    cours = get_object_or_404(Cours, pk=pk)
    form = CoursForm(request.POST or None, instance=cours)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Cours modifié.")
        return redirect('personnel:gestion_cours')
    return render(request, 'personnel/cours/form.html', {'form': form, 'objet': cours, 'titre': 'Modifier le cours'})


@role_required('DIRECTION')
def supprimer_cours(request, pk):
    cours = get_object_or_404(Cours, pk=pk)
    if request.method == 'POST':
        cours.delete()
        messages.success(request, "Cours supprimé.")
        return redirect('personnel:gestion_cours')
    return render(request, 'personnel/confirm_suppression.html', {'objet': cours, 'type': 'le cours'})


# ─── Comptes Personne ─────────────────────────────────────────────────────────

@role_required('DIRECTION')
def gestion_comptes(request):
    qs = Personne.objects.prefetch_related('groups').order_by('nom', 'prenom')
    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'personnel/comptes/liste.html', {'page_obj': page_obj})


@role_required('DIRECTION')
def creer_compte(request):
    role_form = RoleChoiceForm(request.POST or None)
    personne_form = PersonneBaseForm(request.POST or None, request.FILES or None)
    sub_forms = {
        'personnel':  PersonnelSubForm(request.POST or None),
        'enseignant': EnseignantSubForm(request.POST or None),
        'eleve':      EleveSubForm(request.POST or None),
        'parent':     ParentSubForm(request.POST or None),
    }

    if request.method == 'POST':
        role = request.POST.get('role')
        sub_form = sub_forms.get(role)
        if personne_form.is_valid() and role_form.is_valid() and sub_form and sub_form.is_valid():
            personne = personne_form.save()
            _assigner_role(personne, role, sub_form)
            HistoriqueActions.objects.create(
                auteur=request.user, action=f'Création compte ({role})',
                table_cible='personne', id_enreg=personne.pk,
            )
            messages.success(request, f"Compte de {personne.get_full_name()} créé avec succès.")
            return redirect('personnel:gestion_comptes')

    return render(request, 'personnel/comptes/form.html', {
        'personne_form': personne_form, 'role_form': role_form,
        'sub_forms': sub_forms, 'titre': 'Créer un compte',
    })


@role_required('DIRECTION')
def modifier_compte(request, pk):
    personne = get_object_or_404(Personne, pk=pk)
    role = personne.get_role()
    personne_form = PersonneBaseForm(request.POST or None, request.FILES or None, instance=personne)

    sub_instance_map = {
        'personnel': Personnel, 'enseignant': Enseignant,
        'eleve': Eleve, 'parent': Parent,
    }
    sub_form_map = {
        'personnel': PersonnelSubForm, 'enseignant': EnseignantSubForm,
        'eleve': EleveSubForm, 'parent': ParentSubForm,
    }
    sub_instance = sub_instance_map[role].objects.filter(pk=personne.pk).first() if role else None
    SubFormClass = sub_form_map.get(role)
    sub_form = SubFormClass(request.POST or None, instance=sub_instance) if SubFormClass else None

    if request.method == 'POST' and personne_form.is_valid():
        if sub_form is None or sub_form.is_valid():
            personne_form.save()
            if sub_form:
                sub = sub_form.save(commit=False)
                sub.personne = personne
                sub.save()
            messages.success(request, "Compte modifié.")
            return redirect('personnel:gestion_comptes')

    return render(request, 'personnel/comptes/form_modifier.html', {
        'personne_form': personne_form, 'sub_form': sub_form,
        'personne': personne, 'role': role,
        'titre': f'Modifier — {personne.get_full_name()}',
    })


@role_required('DIRECTION')
def toggle_actif_compte(request, pk):
    personne = get_object_or_404(Personne, pk=pk)
    if request.method == 'POST':
        personne.actif = not personne.actif
        personne.save()
        etat = "activé" if personne.actif else "désactivé"
        messages.success(request, f"Compte de {personne.get_full_name()} {etat}.")
        HistoriqueActions.objects.create(
            auteur=request.user, action=f'Compte {etat}',
            table_cible='personne', id_enreg=personne.pk,
        )
    return redirect('personnel:gestion_comptes')


def _assigner_role(personne, role, sub_form):
    import uuid
    sub = sub_form.save(commit=False)
    sub.personne = personne
    if role == 'eleve' and not getattr(sub, 'matricule', None):
        sub.matricule = f"ELV-{uuid.uuid4().hex[:8].upper()}"
    sub.save()
    if role == 'personnel':
        fonction = sub_form.cleaned_data.get('fonction', '').upper()
        group_name = fonction if fonction in ('DIRECTION', 'ADMINISTRATION', 'SCOLARITE', 'FINANCES') else 'ADMINISTRATION'
    else:
        group_name = {'enseignant': 'ENSEIGNANT', 'eleve': 'ELEVE', 'parent': 'PARENT'}[role]
    group = Group.objects.get(name=group_name)
    personne.groups.add(group)


# ─── Demandes d'inscription (auto-inscrits sans rôle) ─────────────────────────

@role_required('DIRECTION', 'SCOLARITE')
def demandes_inscription(request):
    en_attente = Personne.objects.filter(
        groups__isnull=True, actif=False, is_superuser=False
    ).order_by('-date_creation')
    return render(request, 'personnel/comptes/demandes.html', {'en_attente': en_attente})


@role_required('DIRECTION', 'SCOLARITE')
def attribuer_role(request, pk):
    personne = get_object_or_404(Personne, pk=pk, groups__isnull=True, actif=False)
    ROLES_VALIDES = ('personnel', 'enseignant', 'eleve', 'parent')
    role_selectionne = ''
    sub_form_actif = None
    erreurs = []

    sub_forms = {
        'personnel':  PersonnelSubForm(),
        'enseignant': EnseignantSubForm(),
        'eleve':      EleveSubForm(),
        'parent':     ParentSubForm(),
    }

    if request.method == 'POST':
        role_selectionne = request.POST.get('role', '')
        if role_selectionne not in ROLES_VALIDES:
            erreurs.append("Veuillez sélectionner un rôle valide.")
        else:
            sub_form_actif = {
                'personnel':  PersonnelSubForm(request.POST),
                'enseignant': EnseignantSubForm(request.POST),
                'eleve':      EleveSubForm(request.POST),
                'parent':     ParentSubForm(request.POST),
            }[role_selectionne]
            sub_forms[role_selectionne] = sub_form_actif

            if sub_form_actif.is_valid():
                try:
                    _assigner_role(personne, role_selectionne, sub_form_actif)
                    personne.actif = True
                    personne.save()
                    from core.models import Notification
                    Notification.objects.filter(
                        destinataire=request.user,
                        message__icontains=personne.email,
                    ).update(lu=True)
                    HistoriqueActions.objects.create(
                        auteur=request.user,
                        action=f'Attribution role {role_selectionne} a {personne.get_full_name()}',
                        table_cible='personne', id_enreg=personne.pk,
                    )
                    messages.success(request, f"Role attribue et compte de {personne.get_full_name()} active.")
                    return redirect('personnel:demandes_inscription')
                except Exception as e:
                    erreurs.append(f"Erreur lors de l'enregistrement : {e}")
            else:
                for field, errs in sub_form_actif.errors.items():
                    for err in errs:
                        erreurs.append(f"Champ '{field}' : {err}")

    return render(request, 'personnel/comptes/attribuer_role.html', {
        'personne': personne,
        'sub_forms': sub_forms,
        'role_selectionne': role_selectionne,
        'erreurs': erreurs,
    })


@role_required('DIRECTION', 'SCOLARITE')
def rejeter_inscription(request, pk):
    personne = get_object_or_404(Personne, pk=pk, groups__isnull=True, actif=False)
    if request.method == 'POST':
        nom = personne.get_full_name()
        personne.delete()
        messages.warning(request, f"Demande de {nom} rejetée et supprimée.")
    return redirect('personnel:demandes_inscription')


# ─── Inscriptions ─────────────────────────────────────────────────────────────

@role_required('SCOLARITE', 'DIRECTION')
def gestion_inscriptions(request):
    annee_active = AnneeScolaire.objects.filter(active=True).first()
    inscriptions = Inscription.objects.filter(annee=annee_active).select_related(
        'eleve', 'classe', 'annee'
    ).order_by('classe__nom', 'eleve__nom') if annee_active else Inscription.objects.none()

    # Recherche
    q = request.GET.get('q', '').strip()
    if q:
        inscriptions = inscriptions.filter(
            eleve__nom__icontains=q
        ) | inscriptions.filter(
            eleve__prenom__icontains=q
        ) | inscriptions.filter(
            eleve__eleve__matricule__icontains=q
        )
        inscriptions = inscriptions.distinct().order_by('classe__nom', 'eleve__nom')

    # Filtre classe
    classe_filtre = request.GET.get('classe', '')
    if classe_filtre:
        inscriptions = inscriptions.filter(classe_id=classe_filtre)

    paginator = Paginator(inscriptions, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    form = InscriptionForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        try:
            form.save()
            messages.success(request, "Inscription enregistrée.")
            return redirect('personnel:gestion_inscriptions')
        except IntegrityError:
            messages.error(request, "Cet élève est déjà inscrit pour cette année scolaire.")
    classes = Classe.objects.filter(annee=annee_active).order_by('nom') if annee_active else Classe.objects.none()
    return render(request, 'personnel/inscriptions/liste.html', {
        'page_obj': page_obj, 'form': form, 'annee_active': annee_active,
        'classes': classes, 'q': q, 'classe_filtre': classe_filtre,
    })


@role_required('SCOLARITE', 'DIRECTION')
def changer_statut_inscription(request, pk):
    inscription = get_object_or_404(Inscription, pk=pk)
    if request.method == 'POST':
        nouveau_statut = request.POST.get('statut')
        if nouveau_statut in ('transfere', 'abandonne'):
            inscription.statut = nouveau_statut
            inscription.save()
            messages.success(request, f"Statut mis à jour : {inscription.get_statut_display()}.")
    return redirect('personnel:gestion_inscriptions')


# ─── Absences (validation) ────────────────────────────────────────────────────

@role_required('SCOLARITE', 'DIRECTION')
def validation_absences(request):
    absences = Absence.objects.filter(statut='en_attente').select_related(
        'eleve', 'cours__matiere', 'periode'
    ).order_by('-date_saisie')
    return render(request, 'personnel/absences/liste.html', {'absences': absences})


@role_required('SCOLARITE', 'DIRECTION')
def valider_absence(request, pk):
    absence = get_object_or_404(Absence, pk=pk)
    if request.method == 'POST':
        decision = request.POST.get('decision')
        if decision in ('justifiee', 'non_justifiee'):
            absence.statut = decision
            absence.motif = request.POST.get('motif', '')
            absence.save()
            messages.success(request, "Absence traitée.")
    return redirect('personnel:validation_absences')


# ─── Bulletins ────────────────────────────────────────────────────────────────

@role_required('SCOLARITE', 'DIRECTION')
def gestion_bulletins(request):
    annee = AnneeScolaire.objects.filter(active=True).first()
    periodes_qs = Periode.objects.filter(annee=annee).order_by('date_debut') if annee else []
    periodes_data = [
        {'periode': p, 'nb_bulletins': Bulletin.objects.filter(periode=p).count()}
        for p in periodes_qs
    ]
    return render(request, 'personnel/bulletins/liste.html', {
        'annee': annee,
        'periodes_data': periodes_data,
        'today': date.today(),
    })


@role_required('SCOLARITE', 'DIRECTION')
def generer_bulletins_periode(request, pk):
    periode = get_object_or_404(Periode, pk=pk)
    if periode.date_fin > date.today():
        messages.error(request, "Impossible de générer les bulletins : la période n'est pas encore clôturée.")
        return redirect('personnel:gestion_bulletins')

    annee = periode.annee
    classes = Classe.objects.filter(annee=annee)
    nb_crees = 0

    for classe in classes:
        inscriptions = Inscription.objects.filter(
            classe=classe, annee=annee, statut='actif'
        ).select_related('eleve')

        # Calcul de la moyenne générale par élève
        moyennes = []
        for insc in inscriptions:
            resultats = ResultatMatiere.objects.filter(
                eleve=insc.eleve, cours__classe=classe, periode=periode
            )
            moy = resultats.aggregate(moy=Avg('moyenne'))['moy'] or 0
            moyennes.append((insc.eleve, round(moy, 2)))

        # Calcul du rang (tri décroissant par moyenne)
        moyennes.sort(key=lambda x: x[1], reverse=True)

        for rang, (eleve, moy_gen) in enumerate(moyennes, start=1):
            _, created = Bulletin.objects.get_or_create(
                eleve=eleve,
                periode=periode,
                defaults={
                    'moyenne_generale': moy_gen,
                    'rang': rang,
                    'effectif_classe': len(moyennes),
                }
            )
            if created:
                nb_crees += 1

    HistoriqueActions.objects.create(
        auteur=request.user,
        action=f"Génération bulletins période {periode}",
        table_cible='bulletin',
        id_enreg=0,
    )
    messages.success(request, f"{nb_crees} bulletin(s) généré(s) pour la période « {periode} ».")
    return redirect('personnel:gestion_bulletins')


@role_required('SCOLARITE', 'DIRECTION')
def liste_bulletins_periode(request, pk):
    periode = get_object_or_404(Periode, pk=pk)
    bulletins = Bulletin.objects.filter(periode=periode).select_related(
        'eleve', 'eleve__eleve'
    ).order_by('rang')
    annee_active = AnneeScolaire.objects.filter(active=True).first()
    classes = Classe.objects.filter(annee=annee_active).order_by('nom') if annee_active else Classe.objects.none()
    return render(request, 'personnel/bulletins/detail_periode.html', {
        'periode': periode,
        'bulletins': bulletins,
        'classes': classes,
    })


@role_required('SCOLARITE', 'DIRECTION', 'ENSEIGNANT', 'ELEVE', 'PARENT')
def export_bulletin_pdf(request, bulletin_id):
    bulletin = get_object_or_404(Bulletin, pk=bulletin_id)
    eleve = bulletin.eleve
    periode = bulletin.periode

    # Cloisonnement : élève et parent ne peuvent voir que leurs propres bulletins
    user = request.user
    groupes = user.groups.values_list('name', flat=True)
    if 'ELEVE' in groupes and user != eleve:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied
    if 'PARENT' in groupes:
        from core.models import LienParentEleve
        if not LienParentEleve.objects.filter(parent=user, eleve=eleve).exists():
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied

    resultats = ResultatMatiere.objects.filter(
        eleve=eleve, periode=periode
    ).select_related('cours__matiere').order_by('cours__matiere__nom_matiere')

    absences = Absence.objects.filter(eleve=eleve, periode=periode)
    nb_abs = absences.count()
    nb_just = absences.filter(statut='justifiee').count()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    titre_style = ParagraphStyle('titre', parent=styles['Heading1'],
                                 alignment=TA_CENTER, fontSize=16,
                                 textColor=colors.HexColor('#1a3a5c'))
    sous_titre_style = ParagraphStyle('sous_titre', parent=styles['Normal'],
                                      alignment=TA_CENTER, fontSize=11)
    label_style = ParagraphStyle('label', parent=styles['Normal'],
                                 fontSize=10, textColor=colors.grey)

    story = []

    # En-tête
    story.append(Paragraph("ACADEMIQ", titre_style))
    story.append(Paragraph("Bulletin de notes", sous_titre_style))
    story.append(Spacer(1, 0.5*cm))

    # Infos élève
    info_data = [
        ["Élève :", eleve.get_full_name(), "Période :", str(periode)],
        ["Classe :", str(bulletin.eleve.eleve.get_classe_actuelle() if hasattr(bulletin.eleve, 'eleve') else '—'),
         "Année :", str(periode.annee)],
        ["Rang :", f"{bulletin.rang} / {bulletin.effectif_classe}", "Moy. générale :",
         f"{bulletin.moyenne_generale}/20"],
        ["Absences :", f"{nb_abs} ({nb_just} just.)", "", ""],
    ]
    info_table = Table(info_data, colWidths=[3.5*cm, 5*cm, 3.5*cm, 5*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.HexColor('#f8f9fa'), colors.white]),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#dee2e6')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.5*cm))

    # Tableau des résultats
    story.append(Paragraph("Résultats par matière", ParagraphStyle(
        'section', parent=styles['Heading2'], fontSize=12,
        textColor=colors.HexColor('#1a3a5c'), spaceBefore=6)))

    headers = ["Matière", "Moyenne", "Appréciation"]
    table_data = [headers]
    for r in resultats:
        moy = float(r.moyenne) if r.moyenne else 0
        if moy >= 16:
            apprec = "Très bien"
        elif moy >= 14:
            apprec = "Bien"
        elif moy >= 12:
            apprec = "Assez bien"
        elif moy >= 10:
            apprec = "Passable"
        else:
            apprec = "Insuffisant"
        table_data.append([
            r.cours.matiere.nom_matiere,
            f"{r.moyenne}/20" if r.moyenne else "—",
            apprec,
        ])

    if len(table_data) == 1:
        table_data.append(["Aucun résultat disponible", "", ""])

    res_table = Table(table_data, colWidths=[8*cm, 4*cm, 5*cm])
    res_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(res_table)

    doc.build(story)
    buf.seek(0)
    filename = f"bulletin_{eleve.nom}_{eleve.prenom}_{periode.pk}.pdf"
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


# ─── Emploi du temps ──────────────────────────────────────────────────────────

@role_required('DIRECTION')
def historique_actions(request):
    qs = HistoriqueActions.objects.select_related('auteur').order_by('-date_action')
    # Filtres
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(action__icontains=q) | qs.filter(auteur__nom__icontains=q) | qs.filter(table_cible__icontains=q)
        qs = qs.distinct().order_by('-date_action')
    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'personnel/historique.html', {'page_obj': page_obj, 'q': q})


JOURS_ORDER = ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi']


@role_required('DIRECTION', 'ADMINISTRATION')
def gestion_edt(request):
    annee_active = AnneeScolaire.objects.filter(active=True).first()
    classes = Classe.objects.filter(annee=annee_active).order_by('nom') if annee_active else Classe.objects.none()

    classe_id = request.GET.get('classe')
    periode_id = request.GET.get('periode')
    classe_sel = None
    periode_sel = None
    grille = None

    if annee_active:
        periodes = Periode.objects.filter(annee=annee_active).order_by('date_debut')
    else:
        periodes = Periode.objects.none()

    if classe_id:
        classe_sel = get_object_or_404(Classe, pk=classe_id, annee=annee_active)
    if periode_id:
        periode_sel = get_object_or_404(Periode, pk=periode_id)

    creneaux = EmploiDuTemps.objects.none()
    initial = {}
    if periode_sel:
        initial['periode'] = periode_sel
    form = EmploiDuTempsForm(annee=annee_active, classe=classe_sel, initial=initial)

    if request.method == 'POST':
        form = EmploiDuTempsForm(request.POST, annee=annee_active, classe=classe_sel)
        if form.is_valid():
            try:
                edt = form.save(commit=False)
                edt.full_clean()
                edt.save()
                messages.success(request, "Créneau ajouté avec succès.")
            except Exception as e:
                messages.error(request, str(e))
            return redirect(
                f"{request.path}?classe={classe_id or ''}&periode={periode_id or ''}"
            )
        else:
            for field, errs in form.errors.items():
                for err in errs:
                    messages.error(request, f"{field} : {err}")

    if classe_sel and periode_sel:
        creneaux = EmploiDuTemps.objects.filter(
            cours__classe=classe_sel, periode=periode_sel
        ).select_related('cours__matiere', 'cours__enseignant', 'salle').order_by(
            'jour', 'heure_debut'
        )

    return render(request, 'personnel/edt/liste.html', {
        'annee_active': annee_active,
        'classes': classes,
        'periodes': periodes,
        'classe_sel': classe_sel,
        'periode_sel': periode_sel,
        'creneaux': creneaux,
        'form': form,
        'classe_id': classe_id or '',
        'periode_id': periode_id or '',
    })


@role_required('DIRECTION', 'ADMINISTRATION')
def supprimer_creneau(request, pk):
    creneau = get_object_or_404(EmploiDuTemps, pk=pk)
    classe_id = creneau.cours.classe_id
    periode_id = creneau.periode_id
    if request.method == 'POST':
        creneau.delete()
        messages.success(request, "Créneau supprimé.")
    from django.urls import reverse
    return redirect(f"{reverse('personnel:gestion_edt')}?classe={classe_id}&periode={periode_id}")


# ─── Export Excel ─────────────────────────────────────────────────────────────

@role_required('DIRECTION', 'SCOLARITE', 'ADMINISTRATION')
def export_liste_classe_excel(request, classe_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    classe = get_object_or_404(Classe, pk=classe_id)
    annee = AnneeScolaire.objects.filter(active=True).first()
    inscriptions = Inscription.objects.filter(
        classe=classe, annee=annee, statut='actif'
    ).select_related('eleve').order_by('eleve__nom', 'eleve__prenom')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste de classe"

    navy = "0B1D3A"
    gold = "D97706"
    light = "F3F4F6"

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    title_font   = Font(bold=True, color=navy, size=14)
    navy_fill    = PatternFill("solid", fgColor=navy)
    gold_fill    = PatternFill("solid", fgColor=gold)
    light_fill   = PatternFill("solid", fgColor="F9FAFB")
    center       = Alignment(horizontal="center", vertical="center")
    thin         = Side(style="thin", color="D1D5DB")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titre
    ws.merge_cells("A1:F1")
    ws["A1"] = f"ACADEMIQ — Liste de classe : {classe.nom}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Année scolaire : {annee.libelle if annee else '—'}"
    ws["A2"].alignment = center
    ws["A2"].font = Font(italic=True, color="6B7280")

    ws.append([])  # ligne vide

    # En-têtes
    headers = ["N°", "Nom", "Prénom", "Matricule", "Date de naissance", "Statut"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = center
        cell.border = border

    # Données
    for i, insc in enumerate(inscriptions, 1):
        e = insc.eleve
        row = [
            i,
            e.nom,
            e.prenom,
            e.eleve.matricule if hasattr(e, 'eleve') else "—",
            e.eleve.date_naissance.strftime("%d/%m/%Y") if hasattr(e, 'eleve') and e.eleve.date_naissance else "—",
            insc.get_statut_display(),
        ]
        ws.append(row)
        row_num = ws.max_row
        fill = light_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, 7):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Largeurs colonnes
    for col, width in zip(range(1, 7), [6, 20, 20, 15, 18, 12]):
        ws.column_dimensions[get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="liste_{classe.nom}.xlsx"'
    wb.save(response)
    return response


@role_required('DIRECTION', 'SCOLARITE', 'ADMINISTRATION')
def export_notes_excel(request, classe_id, periode_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    classe  = get_object_or_404(Classe, pk=classe_id)
    periode = get_object_or_404(Periode, pk=periode_id)
    annee   = AnneeScolaire.objects.filter(active=True).first()

    inscriptions = Inscription.objects.filter(
        classe=classe, annee=annee, statut='actif'
    ).select_related('eleve').order_by('eleve__nom')

    cours_list = Cours.objects.filter(
        classe=classe, annee=annee
    ).select_related('matiere').order_by('matiere__nom_matiere')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relevé de notes"

    navy_fill   = PatternFill("solid", fgColor="0B1D3A")
    gold_fill   = PatternFill("solid", fgColor="D97706")
    green_fill  = PatternFill("solid", fgColor="DCFCE7")
    red_fill    = PatternFill("solid", fgColor="FEE2E2")
    light_fill  = PatternFill("solid", fgColor="F9FAFB")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin        = Side(style="thin", color="D1D5DB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center")

    # Titre
    nb_cols = 2 + len(cours_list) + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
    ws.cell(1, 1).value = f"ACADEMIQ — Relevé de notes : {classe.nom} / {periode.nom}"
    ws.cell(1, 1).font = Font(bold=True, color="0B1D3A", size=13)
    ws.cell(1, 1).alignment = center
    ws.row_dimensions[1].height = 26
    ws.append([])

    # En-têtes
    row_h = ["N°", "Élève"] + [c.matiere.nom_matiere[:15] for c in cours_list] + ["Moy. Gén."]
    ws.append(row_h)
    for col in range(1, len(row_h) + 1):
        cell = ws.cell(3, col)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = center
        cell.border = border

    # Données
    for i, insc in enumerate(inscriptions, 1):
        eleve = insc.eleve
        row_data = [i, eleve.get_full_name()]
        total_pts, total_coef = 0, 0
        for cours in cours_list:
            res = ResultatMatiere.objects.filter(
                eleve=eleve, cours=cours, periode=periode
            ).first()
            moy = float(res.moyenne) if res and res.moyenne else None
            row_data.append(f"{moy:.2f}" if moy is not None else "—")
            if moy is not None:
                coef = float(cours.matiere.coefficient)
                total_pts  += moy * coef
                total_coef += coef

        moy_gen = round(total_pts / total_coef, 2) if total_coef else None
        row_data.append(f"{moy_gen:.2f}" if moy_gen is not None else "—")
        ws.append(row_data)

        row_num = ws.max_row
        alt = light_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row_num, col)
            cell.alignment = center
            cell.border = border
            if col == len(row_data) and moy_gen is not None:
                cell.fill = green_fill if moy_gen >= 10 else red_fill
                cell.font = Font(bold=True)
            else:
                cell.fill = alt

    # Largeurs
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    for col in range(3, nb_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="notes_{classe.nom}_{periode.nom}.xlsx"'
    wb.save(response)
    return response


# ─── Attestation de scolarité PDF ─────────────────────────────────────────────

@role_required('DIRECTION', 'SCOLARITE', 'ADMINISTRATION')
def attestation_scolarite_pdf(request, inscription_id):
    from datetime import date as today_date
    inscription = get_object_or_404(Inscription, pk=inscription_id, statut='actif')
    eleve = inscription.eleve
    classe = inscription.classe
    annee  = inscription.annee

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2.5*cm, rightMargin=2.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                 alignment=TA_CENTER, fontSize=16,
                                 textColor=colors.HexColor('#0B1D3A'), spaceAfter=4)
    center_style = ParagraphStyle('center', parent=styles['Normal'],
                                  alignment=TA_CENTER, fontSize=11)
    body_style   = ParagraphStyle('body', parent=styles['Normal'],
                                  fontSize=12, leading=20)
    label_style  = ParagraphStyle('label', parent=styles['Normal'],
                                  fontSize=10, textColor=colors.grey)

    story = []

    # En-tête établissement
    story.append(Paragraph("ACADEMIQ", title_style))
    story.append(Paragraph("Établissement d'enseignement secondaire", center_style))
    story.append(Spacer(1, 0.3*cm))

    # Ligne dorée
    sep = Table([['']], colWidths=[16*cm])
    sep.setStyle(TableStyle([('LINEBELOW', (0,0), (-1,-1), 2, colors.HexColor('#D97706'))]))
    story.append(sep)
    story.append(Spacer(1, 0.8*cm))

    # Titre du document
    story.append(Paragraph("ATTESTATION DE SCOLARITÉ", ParagraphStyle(
        'attest', parent=styles['Heading1'], alignment=TA_CENTER,
        fontSize=18, textColor=colors.HexColor('#0B1D3A'),
        borderPad=10, spaceAfter=6,
    )))
    story.append(Paragraph(f"Année scolaire {annee.libelle}", center_style))
    story.append(Spacer(1, 1*cm))

    # Corps du texte
    texte = (
        f"Le Directeur de l'établissement ACADEMIQ certifie que :"
    )
    story.append(Paragraph(texte, body_style))
    story.append(Spacer(1, 0.5*cm))

    # Infos élève dans un tableau
    data_eleve = [
        ["Nom et Prénom :", eleve.get_full_name()],
        ["Matricule :",     eleve.eleve.matricule if hasattr(eleve, 'eleve') else "—"],
        ["Date de naissance :", eleve.eleve.date_naissance.strftime("%d/%m/%Y")
         if hasattr(eleve, 'eleve') and eleve.eleve.date_naissance else "—"],
        ["Classe :", classe.nom],
        ["Année scolaire :", annee.libelle],
    ]
    t = Table(data_eleve, colWidths=[5*cm, 11*cm])
    t.setStyle(TableStyle([
        ('FONTNAME',  (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE',  (0, 0), (-1, -1), 12),
        ('TOPPADDING',  (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1),
         [colors.HexColor('#F9FAFB'), colors.white]),
        ('BOX',   (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('GRID',  (0, 0), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.8*cm))

    suite = (
        f"est régulièrement inscrit(e) dans notre établissement pour l'année scolaire "
        f"<b>{annee.libelle}</b>, en classe de <b>{classe.nom}</b>."
    )
    story.append(Paragraph(suite, body_style))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        "La présente attestation est délivrée pour servir et valoir ce que de droit.",
        body_style
    ))
    story.append(Spacer(1, 1.5*cm))

    # Signature
    date_str = today_date.today().strftime("%d/%m/%Y")
    sig_data = [["Fait le :", date_str, "Le Directeur"]]
    sig_table = Table(sig_data, colWidths=[3*cm, 7*cm, 6*cm])
    sig_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (2, 0), (2, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
    ]))
    story.append(sig_table)
    story.append(Spacer(1, 2*cm))
    story.append(Paragraph("_" * 35, ParagraphStyle('sig', parent=styles['Normal'],
                            alignment=TA_CENTER, fontSize=10, textColor=colors.grey)))

    doc.build(story)
    buf.seek(0)
    nom_fichier = f"attestation_{eleve.nom}_{eleve.prenom}.pdf"
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{nom_fichier}"'
    return response


# ─── Messagerie interne ───────────────────────────────────────────────────────

@role_required(*ROLES_ADMIN, 'ENSEIGNANT', 'ELEVE', 'PARENT')
def messagerie(request):
    recus  = Message.objects.filter(destinataire=request.user).order_by('-date_envoi')
    envoyes = Message.objects.filter(expediteur=request.user).order_by('-date_envoi')
    nb_non_lus = recus.filter(lu=False).count()
    return render(request, 'personnel/messagerie/liste.html', {
        'recus': recus[:20], 'envoyes': envoyes[:20], 'nb_non_lus': nb_non_lus,
    })


@role_required(*ROLES_ADMIN, 'ENSEIGNANT', 'ELEVE', 'PARENT')
def envoyer_message(request):
    if request.method == 'POST':
        dest_id = request.POST.get('destinataire')
        sujet   = request.POST.get('sujet', '').strip()
        corps   = request.POST.get('corps', '').strip()
        if dest_id and sujet and corps:
            try:
                dest = Personne.objects.get(pk=dest_id, actif=True)
                Message.objects.create(
                    expediteur=request.user, destinataire=dest,
                    sujet=sujet, corps=corps
                )
                messages.success(request, f"Message envoyé à {dest.get_full_name()}.")
            except Personne.DoesNotExist:
                messages.error(request, "Destinataire introuvable.")
        else:
            messages.error(request, "Veuillez remplir tous les champs.")
        return redirect('personnel:messagerie')

    destinataires = Personne.objects.filter(actif=True).exclude(pk=request.user.pk).order_by('nom')
    return render(request, 'personnel/messagerie/nouveau.html', {'destinataires': destinataires})


@role_required(*ROLES_ADMIN, 'ENSEIGNANT', 'ELEVE', 'PARENT')
def lire_message(request, pk):
    from django.db.models import Q
    msg = Message.objects.filter(
        Q(destinataire=request.user) | Q(expediteur=request.user), pk=pk
    ).first()
    if not msg:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied
    if msg.destinataire == request.user and not msg.lu:
        msg.lu = True
        msg.save(update_fields=['lu'])
    return render(request, 'personnel/messagerie/detail.html', {'msg': msg})


# ─── Annonces générales ───────────────────────────────────────────────────────

@role_required('DIRECTION', 'ADMINISTRATION')
def gestion_annonces(request):
    annee_active = AnneeScolaire.objects.filter(active=True).first()
    annonces = Notification.objects.filter(type_notif='annonce', classe__isnull=False).order_by('-id')

    if request.method == 'POST':
        classe_id = request.POST.get('classe')
        msg_text  = request.POST.get('message', '').strip()
        if classe_id and msg_text:
            try:
                cls = Classe.objects.get(pk=classe_id)
                Notification.objects.create(
                    classe=cls, message=msg_text, type_notif='annonce'
                )
                messages.success(request, f"Annonce envoyée à la classe {cls.nom}.")
            except Classe.DoesNotExist:
                messages.error(request, "Classe introuvable.")
        else:
            messages.error(request, "Message et classe requis.")
        return redirect('personnel:gestion_annonces')

    classes = Classe.objects.filter(annee=annee_active).order_by('nom') if annee_active else Classe.objects.none()
    return render(request, 'personnel/annonces.html', {
        'annonces': annonces, 'classes': classes,
    })


# ─── Module Finances ──────────────────────────────────────────────────────────

@role_required('DIRECTION', 'FINANCES', 'ADMINISTRATION')
def gestion_finances(request):
    annee_active = AnneeScolaire.objects.filter(active=True).first()
    frais_qs = FraisScolarite.objects.filter(annee=annee_active).select_related(
        'eleve', 'annee'
    ).order_by('statut', 'eleve__nom') if annee_active else FraisScolarite.objects.none()

    # Filtres
    statut_filtre = request.GET.get('statut', '')
    if statut_filtre:
        frais_qs = frais_qs.filter(statut=statut_filtre)

    # Résumé
    from django.db.models import Sum
    total_du   = frais_qs.aggregate(t=Sum('montant_du'))['t'] or 0
    total_paye = frais_qs.aggregate(t=Sum('montant_paye'))['t'] or 0

    paginator = Paginator(frais_qs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))

    eleves_actifs = Personne.objects.filter(
        groups__name='ELEVE', actif=True
    ).order_by('nom', 'prenom')

    return render(request, 'personnel/finances/liste.html', {
        'page_obj': page_obj, 'annee_active': annee_active,
        'total_du': total_du, 'total_paye': total_paye,
        'reste': total_du - total_paye, 'statut_filtre': statut_filtre,
        'eleves_actifs': eleves_actifs,
    })


@role_required('DIRECTION', 'FINANCES', 'ADMINISTRATION')
def creer_frais(request):
    if request.method == 'POST':
        eleve_id   = request.POST.get('eleve')
        montant_du = request.POST.get('montant_du', '0')
        echeance   = request.POST.get('date_echeance') or None
        annee      = AnneeScolaire.objects.filter(active=True).first()
        try:
            eleve = Personne.objects.get(pk=eleve_id, eleve__isnull=False)
            FraisScolarite.objects.create(
                eleve=eleve, annee=annee,
                montant_du=montant_du, date_echeance=echeance,
            )
            messages.success(request, "Frais créés avec succès.")
        except Exception as e:
            messages.error(request, str(e))
    return redirect('personnel:gestion_finances')


@role_required('DIRECTION', 'FINANCES', 'ADMINISTRATION')
def enregistrer_paiement(request, frais_id):
    frais = get_object_or_404(FraisScolarite, pk=frais_id)
    if request.method == 'POST':
        montant  = request.POST.get('montant', '0')
        date_p   = request.POST.get('date_paiement')
        mode     = request.POST.get('mode', 'especes')
        try:
            Paiement.objects.create(
                frais=frais, montant=montant,
                date_paiement=date_p, mode=mode,
                saisi_par=request.user,
            )
            messages.success(request, "Paiement enregistré.")
            HistoriqueActions.objects.create(
                auteur=request.user, action='Paiement enregistré',
                table_cible='paiement', id_enreg=frais.pk,
                details=f"Montant : {montant} FCFA"
            )
        except Exception as e:
            messages.error(request, str(e))
    return redirect('personnel:gestion_finances')


@role_required('DIRECTION', 'FINANCES', 'ADMINISTRATION')
def recu_paiement_pdf(request, paiement_id):
    paiement = get_object_or_404(Paiement, pk=paiement_id)
    frais    = paiement.frais
    eleve    = frais.eleve

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2.5*cm, rightMargin=2.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    title_s = ParagraphStyle('t', parent=styles['Heading1'], alignment=TA_CENTER,
                              fontSize=16, textColor=colors.HexColor('#0B1D3A'))
    center_s = ParagraphStyle('c', parent=styles['Normal'], alignment=TA_CENTER, fontSize=11)

    story.append(Paragraph("ACADEMIQ", title_s))
    story.append(Paragraph("Reçu de paiement des frais de scolarité", center_s))
    story.append(Spacer(1, 0.4*cm))

    sep = Table([['']], colWidths=[16*cm])
    sep.setStyle(TableStyle([('LINEBELOW', (0,0), (-1,-1), 2, colors.HexColor('#D97706'))]))
    story.append(sep)
    story.append(Spacer(1, 0.8*cm))

    data = [
        ["N° Reçu :", paiement.recu_numero],
        ["Date :", paiement.date_paiement.strftime("%d/%m/%Y")],
        ["Élève :", eleve.get_full_name()],
        ["Année scolaire :", str(frais.annee)],
        ["Montant payé :", f"{paiement.montant:,.0f} FCFA"],
        ["Mode de paiement :", paiement.get_mode_display()],
        ["Total versé :", f"{frais.montant_paye:,.0f} / {frais.montant_du:,.0f} FCFA"],
        ["Reste à payer :", f"{frais.reste_a_payer:,.0f} FCFA"],
    ]
    t = Table(data, colWidths=[5*cm, 11*cm])
    t.setStyle(TableStyle([
        ('FONTNAME',  (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE',  (0, 0), (-1, -1), 12),
        ('TOPPADDING',    (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.HexColor('#F9FAFB'), colors.white]),
        ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#E5E7EB')),
        ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#DCFCE7')),
        ('FONTNAME', (1, 4), (1, 4), 'Helvetica-Bold'),
    ]))
    story.append(t)
    story.append(Spacer(1, 1.5*cm))
    story.append(Paragraph("Cachet et signature", ParagraphStyle(
        'sig', parent=styles['Normal'], alignment=TA_CENTER,
        fontSize=10, textColor=colors.grey)))

    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="recu_{paiement.recu_numero}.pdf"'
    return response


# ─── Calendrier scolaire ──────────────────────────────────────────────────────

@role_required(*ROLES_ADMIN, 'ENSEIGNANT', 'ELEVE', 'PARENT')
def calendrier_scolaire(request):
    import json
    annee_active = AnneeScolaire.objects.filter(active=True).first()
    evenements = EvenementScolaire.objects.filter(
        annee=annee_active
    ).order_by('date_debut') if annee_active else EvenementScolaire.objects.none()

    couleurs = {
        'vacances': '#D97706', 'examen': '#dc2626',
        'reunion': '#2563eb', 'ferie': '#16a34a', 'autre': '#6b7280',
    }
    events_json = json.dumps([{
        'title': e.titre,
        'start': e.date_debut.isoformat(),
        'end':   e.date_fin.isoformat(),
        'color': couleurs.get(e.type_event, '#6b7280'),
        'description': e.description,
    } for e in evenements])

    return render(request, 'personnel/calendrier.html', {
        'evenements': evenements, 'annee_active': annee_active,
        'events_json': events_json,
    })


@role_required('DIRECTION', 'ADMINISTRATION')
def creer_evenement(request):
    if request.method == 'POST':
        titre    = request.POST.get('titre', '').strip()
        type_ev  = request.POST.get('type_event', 'autre')
        date_deb = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        desc     = request.POST.get('description', '')
        annee    = AnneeScolaire.objects.filter(active=True).first()
        if titre and date_deb and date_fin:
            EvenementScolaire.objects.create(
                titre=titre, type_event=type_ev,
                date_debut=date_deb, date_fin=date_fin,
                description=desc, annee=annee, createur=request.user,
            )
            messages.success(request, "Événement ajouté au calendrier.")
        else:
            messages.error(request, "Titre et dates requis.")
    return redirect('personnel:calendrier_scolaire')


@role_required('DIRECTION', 'ADMINISTRATION')
def supprimer_evenement(request, pk):
    ev = get_object_or_404(EvenementScolaire, pk=pk)
    if request.method == 'POST':
        ev.delete()
        messages.success(request, "Événement supprimé.")
    return redirect('personnel:calendrier_scolaire')
